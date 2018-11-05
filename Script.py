import sys
import pandas as pd
import openpyxl
import csv
import datetime
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


#EXAMPLE USAGE: python.exe Script.py export export(1)
filenames = []
#1. open the two files or one
if(len(sys.argv) == 2):
    first_report = 'C:\\Users\\f5mgs\\Downloads\\' + sys.argv[1] + '.csv'
    filenames.append(first_report)

if(len(sys.argv) == 3):
    first_report = 'C:\\Users\\f5mgs\\Downloads\\' + sys.argv[1] + '.csv'
    second_report = 'C:\\Users\\f5mgs\\Downloads\\' + sys.argv[2] + '.csv'
    filenames.append(first_report)
    filenames.append(second_report)

outfile = 'C:\\Users\\f5mgs\\Desktop\\Co-op\\Change Meetings\\combined.csv'

#2. combine the files
if(len(filenames) == 2):

    a = pd.read_csv(first_report)
    b = pd.read_csv(second_report)
    frames = [a,b]
    result = pd.concat(frames)
    result.to_csv(outfile, index=False)
else:
    a = pd.read_csv(first_report)
    frames = [a]
    result = pd.concat(frames)
    result.to_csv(outfile, index=False)

wb = openpyxl.Workbook()
ws = wb.active

with open('C:\\Users\\f5mgs\\Desktop\\Co-op\\Change Meetings\\combined.csv', 'rU') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

new_outfile = 'C:\\Users\\f5mgs\\Desktop\\Co-op\\Change Meetings\\CM_20181001.xlsx'

#3. save to a new workbook in .xlsx
wb.save(new_outfile)

#4. then open it back up as an .xlsx
wb = openpyxl.load_workbook(new_outfile)
sheet = wb.active

#5. delete the rows with change managment date of not today's date
sheet_size = sheet.max_row
today = datetime.datetime.today().date().strftime('%m/%d/%Y')
shifted = True

for loop in range(20):
    for ro in range(sheet_size + 1):
        if ro == 0 or ro == 1:
            continue
        cm_date = sheet.cell(row=ro, column=9)
        if cm_date.value == None:
            sheet.delete_rows(ro,1)
            continue
        cm_date = cm_date.value.split(' ')
        #take the split up cell value and cylce thru if today's date is a match than keep it.
        found = False
        for split in cm_date:
            print split
            print today
            if (split.find(today) > -1):
                print "found"
                found = True
                break
        if(found == False):
            sheet.delete_rows(ro,1)

#6 . highlight the rows with change scheuled downtime start or  ad hoc director to true
yelFill = openpyxl.styles.PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')

sheet_size = sheet.max_row
for ro in range(sheet_size + 1):
    if ro == 0 or ro == 1:
        continue
    if (sheet.cell(row=ro, column=8).value == 'True') or (sheet.cell(row=ro, column=5).value != None):
        sheet.cell(row=ro, column=8).value = 'Yes'
        for r in range(sheet.max_column + 1):
            if r == 0:
                continue
            sheet.cell(row=ro, column=r).fill = yelFill

#7. Do borders on every cell and align left
side = Side(border_style='thin', color="FF000000")
column_size = sheet.max_column
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for i in range(sheet_size + 1):
    if i == 0:
        continue
    for j in range(column_size + 1):
        if j == 0:
            continue
        cell = sheet.cell(row=i, column=j)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left')

#8. Normalize column widths
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['B'].width = 50
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
    sheet.column_dimensions[col].width = 30

#9. Change wording on column header
sheet.cell(row=1, column=3).value = 'Implementation Start'
sheet.cell(row=1, column=4).value = 'Implementation End'
sheet.cell(row=1, column=5).value = 'Downtime Start'
sheet.cell(row=1, column=6).value = 'Downtime End'
sheet.cell(row=1, column=7).value = 'Coordinator'

#10. hide the columns that are not needed.
for col in ['J', 'K', 'L']:
    sheet.column_dimensions[col].hidden = True

#11. save and close the file.
wb.save(new_outfile)

print 'program is done.'